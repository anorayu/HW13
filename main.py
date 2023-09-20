import pandas as pd

df1 = pd.read_excel('1111.xlsx')
df2 = pd.read_excel('2222.xlsx')
df3 = pd.read_excel('3333.xlsx')
df1_sorted = df1.sort_values(by='column_name', ascending=False
df1_sorted.to_excel('output.xlsx', index=False)
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# Загрузка файла
wb = load_workbook('output.xlsx')

# Выбор активного листа
ws = wb.active

# Изменение шрифта
font = Font(name='Arial', bold=True)
ws['A1'].font = font

# Добавление границ
border = Border(left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin'))
ws['A1'].border = border

# Сохранение изменений
wb.save('output.xlsx')
